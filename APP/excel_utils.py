from openpyxl import load_workbook
import datetime

def atualizar_planilha(resultado):
    caminho = "___________________"
    wb = load_workbook(caminho, keep_vba=True)
    ws = wb.active

    nova_linha = ws.max_row + 1
    ws[f"A{nova_linha}"] = resultado.operador
    ws[f"B{nova_linha}"] = str(resultado.dados)
    ws[f"C{nova_linha}"] = resultado.observacoes
    ws[f"D{nova_linha}"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    wb.save(caminho)
